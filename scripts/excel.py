import json
import sys
import re
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Excel által tiltott láthatatlan vezérlőkarakterek
ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def safe_cell(value):
    """
    Minden cellába kerülő értéket Excel-biztossá alakít.
    Ez javítja az IllegalCharacterError hibát is.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    else:
        value = str(value)

    value = ILLEGAL_CHARS.sub("", value)
    value = value.replace("\u2028", " ").replace("\u2029", " ")

    # Excel formula injection elkerülése naplózott támadói inputoknál
    if value.startswith(("=", "+", "-", "@")):
        value = "'" + value

    # Excel cella maximum karakterhossz
    if len(value) > 32767:
        value = value[:32760] + "..."

    return value


def load_cowrie_json(path: Path):
    events = []

    with path.open("r", encoding="utf-8", errors="ignore") as file:
        for line_number, line in enumerate(file, start=1):
            line = line.strip()

            if not line:
                continue

            try:
                event = json.loads(line)
                events.append(event)
            except json.JSONDecodeError:
                print(f"Hibás JSON sor kihagyva: {line_number}")

    return events


def style_header(sheet):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            value = cell.value
            if value is None:
                continue

            value = str(value)
            if len(value) > max_length:
                max_length = len(value)

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def finish_sheet(sheet):
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        style_header(sheet)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    autosize_columns(sheet)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_table(sheet, headers, rows):
    sheet.append([safe_cell(header) for header in headers])

    for row in rows:
        sheet.append([safe_cell(row.get(header, "")) for header in headers])

    finish_sheet(sheet)


def create_excel(events, output_path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # -------------------------
    # Összegzés
    # -------------------------
    summary = workbook.create_sheet("Osszegzes")

    event_counter = Counter(event.get("eventid", "unknown") for event in events)
    ip_counter = Counter(event.get("src_ip", "") for event in events if event.get("src_ip"))
    username_counter = Counter(event.get("username", "") for event in events if event.get("username"))
    password_counter = Counter(event.get("password", "") for event in events if event.get("password"))

    commands = [
        event.get("input", "")
        for event in events
        if event.get("eventid") == "cowrie.command.input"
    ]
    command_counter = Counter(commands)

    sessions = set(event.get("session") for event in events if event.get("session"))

    summary.append(["Mutató", "Érték"])
    summary.append(["Összes esemény", len(events)])
    summary.append(["Egyedi sessionök száma", len(sessions)])
    summary.append(["Egyedi forrás IP-k száma", len(ip_counter)])
    summary.append(["Parancs események száma", len(commands)])
    summary.append(["Sikertelen login próbák száma", event_counter.get("cowrie.login.failed", 0)])
    summary.append(["Sikeres login próbák száma", event_counter.get("cowrie.login.success", 0)])

    row = summary.max_row + 3
    summary.cell(row=row, column=1, value="Leggyakoribb eseménytípusok")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value="eventid")
    summary.cell(row=row, column=2, value="darab")
    row += 1

    for eventid, count in event_counter.most_common(20):
        summary.cell(row=row, column=1, value=safe_cell(eventid))
        summary.cell(row=row, column=2, value=count)
        row += 1

    row += 2
    summary.cell(row=row, column=1, value="Leggyakoribb forrás IP-k")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value="IP")
    summary.cell(row=row, column=2, value="darab")
    row += 1

    for ip, count in ip_counter.most_common(20):
        summary.cell(row=row, column=1, value=safe_cell(ip))
        summary.cell(row=row, column=2, value=count)
        row += 1

    row += 2
    summary.cell(row=row, column=1, value="Leggyakoribb parancsok")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value="parancs")
    summary.cell(row=row, column=2, value="darab")
    row += 1

    for command, count in command_counter.most_common(30):
        summary.cell(row=row, column=1, value=safe_cell(command))
        summary.cell(row=row, column=2, value=count)
        row += 1

    row += 2
    summary.cell(row=row, column=1, value="Leggyakoribb felhasználónevek")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value="felhasználónév")
    summary.cell(row=row, column=2, value="darab")
    row += 1

    for username, count in username_counter.most_common(20):
        summary.cell(row=row, column=1, value=safe_cell(username))
        summary.cell(row=row, column=2, value=count)
        row += 1

    row += 2
    summary.cell(row=row, column=1, value="Leggyakoribb jelszavak")
    summary.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    summary.cell(row=row, column=1, value="jelszó")
    summary.cell(row=row, column=2, value="darab")
    row += 1

    for password, count in password_counter.most_common(20):
        summary.cell(row=row, column=1, value=safe_cell(password))
        summary.cell(row=row, column=2, value=count)
        row += 1

    finish_sheet(summary)

    # -------------------------
    # Login próbálkozások
    # -------------------------
    login_events = []
    for event in events:
        if event.get("eventid") in ("cowrie.login.failed", "cowrie.login.success"):
            login_events.append({
                "timestamp": event.get("timestamp", ""),
                "eventid": event.get("eventid", ""),
                "src_ip": event.get("src_ip", ""),
                "username": event.get("username", ""),
                "password": event.get("password", ""),
                "session": event.get("session", ""),
                "message": event.get("message", ""),
            })

    login_sheet = workbook.create_sheet("Loginok")
    write_table(
        login_sheet,
        ["timestamp", "eventid", "src_ip", "username", "password", "session", "message"],
        login_events
    )

    # -------------------------
    # Parancsok
    # -------------------------
    command_events = []
    for event in events:
        if event.get("eventid") == "cowrie.command.input":
            command_events.append({
                "timestamp": event.get("timestamp", ""),
                "src_ip": event.get("src_ip", ""),
                "session": event.get("session", ""),
                "command": event.get("input", ""),
                "message": event.get("message", ""),
            })

    command_sheet = workbook.create_sheet("Parancsok")
    write_table(
        command_sheet,
        ["timestamp", "src_ip", "session", "command", "message"],
        command_events
    )

    # -------------------------
    # Kapcsolatok
    # -------------------------
    connection_events = []
    for event in events:
        if event.get("eventid") in (
            "cowrie.session.connect",
            "cowrie.session.closed",
            "cowrie.proxy.backend_connected",
            "cowrie.proxy.backend_disconnected",
            "cowrie.client.version",
            "cowrie.client.kex",
        ):
            connection_events.append({
                "timestamp": event.get("timestamp", ""),
                "eventid": event.get("eventid", ""),
                "src_ip": event.get("src_ip", ""),
                "src_port": event.get("src_port", ""),
                "dst_ip": event.get("dst_ip", ""),
                "dst_port": event.get("dst_port", ""),
                "backend_ip": event.get("backend_ip", ""),
                "backend_port": event.get("backend_port", ""),
                "session": event.get("session", ""),
                "message": event.get("message", ""),
            })

    connection_sheet = workbook.create_sheet("Kapcsolatok")
    write_table(
        connection_sheet,
        [
            "timestamp",
            "eventid",
            "src_ip",
            "src_port",
            "dst_ip",
            "dst_port",
            "backend_ip",
            "backend_port",
            "session",
            "message",
        ],
        connection_events
    )

    # -------------------------
    # Minden esemény
    # -------------------------
    preferred_keys = [
        "timestamp",
        "eventid",
        "src_ip",
        "src_port",
        "dst_ip",
        "dst_port",
        "username",
        "password",
        "input",
        "session",
        "message",
        "sensor",
        "protocol",
        "backend_ip",
        "backend_port",
        "local_ip",
        "local_port",
        "duration",
    ]

    all_keys = []
    for key in preferred_keys:
        if any(key in event for event in events):
            all_keys.append(key)

    for event in events:
        for key in event.keys():
            if key not in all_keys:
                all_keys.append(key)

    all_events_sheet = workbook.create_sheet("Minden_esemeny")
    write_table(all_events_sheet, all_keys, events)

    workbook.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Használat: python excel.py cowrie.json")
        sys.exit(1)

    input_path = Path(sys.argv[1])

    if not input_path.exists():
        print(f"Nem található a fájl: {input_path}")
        sys.exit(1)

    events = load_cowrie_json(input_path)

    if not events:
        print("Nem sikerült eseményeket beolvasni.")
        sys.exit(1)

    output_path = input_path.with_name(input_path.stem + "_analysis.xlsx")
    create_excel(events, output_path)

    print(f"Kész: {output_path}")
    print(f"Beolvasott események száma: {len(events)}")


if __name__ == "__main__":
    main()